from flask import Flask, request
import pandas as pd
import os
from openpyxl import load_workbook

app = Flask(__name__)

FILE_NAME = "data.xlsx"

@app.route('/')
def form():
    return '''
    <h2>Data Entry Form</h2>
    <form method="POST" action="/submit">
        Year: <input type="text" name="year"><br><br>
        Component: <input type="text" name="component"><br><br>
        Rec Central: <input type="number" name="rec_c"><br><br>
        Rec State: <input type="number" name="rec_s"><br><br>
        NonRec Central: <input type="number" name="nonrec_c"><br><br>
        NonRec State: <input type="number" name="nonrec_s"><br><br>
        Remarks: <input type="text" name="remarks"><br><br>
        <input type="submit" value="Submit">
    </form>
    '''

@app.route('/submit', methods=['POST'])
def submit():
    data = {
        "Year": request.form['year'],
        "Component": request.form['component'],
        "Rec_Central": request.form['rec_c'],
        "Rec_State": request.form['rec_s'],
        "NonRec_Central": request.form['nonrec_c'],
        "NonRec_State": request.form['nonrec_s'],
        "Remarks": request.form['remarks']
    }

    df = pd.DataFrame([data])

    if not os.path.exists(FILE_NAME):
        df.to_excel(FILE_NAME, index=False)
    else:
        book = load_workbook(FILE_NAME)
        sheet = book.active
        start_row = sheet.max_row

        with pd.ExcelWriter(FILE_NAME, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            df.to_excel(writer, index=False, header=False, startrow=start_row)

    return "Data Saved Successfully!"